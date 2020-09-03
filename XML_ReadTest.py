import xml.etree.ElementTree as ET
import easygui
import openpyxl
import time
from tkinter import filedialog
from tkinter import *
from Util import *  
from datetime import datetime

# Start Program!
print("Select Idex folder")
FolderRoot = Tk()
FolderRoot.dirname = filedialog.askdirectory()
PrintFolderPath(FolderRoot.dirname)

xmlFileList = []

CollectFiles(FolderRoot.dirname, xmlFileList, '.xml')

#Load Excel file
print("Select Excel file")
sExcelPath = easygui.fileopenbox('Select Target Excel File')
# sExcelPath = "C:\\Users\\AXVKI2D\\Desktop\\AVK Test Environment Status.xlsx"
fExcelFile = openpyxl.load_workbook(filename = sExcelPath, read_only = False, keep_vba=True)
MainInfoSheet = fExcelFile.get_sheet_by_name('Audi')

#Set Each Colume in Excel
ColDict = {}
DictionarySetting(MainInfoSheet, ColDict, 3, True)

#Define variables
MUInfo = ComponentInfo()
KombiInfo = ComponentInfo()
GWInfo = ComponentInfo()
cBoxInfo = ComponentInfo()
DBInfo = ComponentInfo()
BoosterInfo = ComponentInfo()
BCM1Info = ComponentInfo()
BCM2Info = ComponentInfo()
HUDInfo =  ComponentInfo()

#Get Data from Xml
for xmlIndex in xmlFileList:
    XmlTree = ET.parse(xmlIndex, parser=ET.XMLParser(encoding='iso-8859-5'))
    XmlRoot = XmlTree.getroot()
    Child = XmlRoot.getchildren()

    #Get VIN Number
    VIN_No = GetIterator(Child, 'Fahrgestellnummer').text      #Get VIN number    
    print("*Start: ", VIN_No)
    VINRow = [1]
    if VIN_No == None:
        VINRow[0] = MainInfoSheet.max_row + 1
    elif GetVINRow(MainInfoSheet, VIN_No, ColDict['VIN'], VINRow) == False:
        MainInfoSheet.cell(VINRow[0], ColDict['VIN']).value = VIN_No
        # fExcelFile.save(filename = sExcelPath) #Save Excel File

    #Get CarLine
    Project = GetIterator(Child, 'UserProjekt').text     
    MainInfoSheet.cell(VINRow[0], ColDict['Car line']).value = Project
    # fExcelFile.save(filename = sExcelPath) #Save Excel File
    print("*Start: ", VIN_No)
    #Get necessary info
    DiagnoseCategory = GetIterator(Child, 'Diagnosebloecke')
    for iterator in DiagnoseCategory:    
        temp = iterator.getchildren()    
        InfoIter = GetIterator(temp, 'Bezeichnung')
        # print(iterator.attrib, InfoIter.text)
                
        if InfoIter.text == "Information Control Unit 1": 
            GetInfoData(temp, MUInfo)   #Get MU Data
            MainInfoSheet.cell(VINRow[0], ColDict["Main Unit"]).value = MUInfo.PartNo
            MainInfoSheet.cell(VINRow[0], ColDict["Main Unit"]+1).value = MUInfo.HWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Main Unit"]+2).value = MUInfo.SWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Main Unit"]+3).value = MUInfo.KeyType

        #Get Booster Data
            SubItemIter = GetIterator(temp, "SubTeilnehmer") 
            if SubItemIter != None:   
                Boostertemp = SubItemIter.getchildren()  
                for BoosterTemp in Boostertemp:
                    BoosterChild = BoosterTemp.getchildren()  
                    BoosterIter = GetIterator(BoosterChild, "SubtName")
                    # print (BoosterIter.text)
                    if BoosterIter.text == "Digital Sound System Control Module 2":
                        GetInfoData(BoosterChild, BoosterInfo)
                        
                        MainInfoSheet.cell(VINRow[0], ColDict["Booster"]).value = BoosterInfo.PartNo
                        MainInfoSheet.cell(VINRow[0], ColDict["Booster"]+1).value = BoosterInfo.HWVersion
                        MainInfoSheet.cell(VINRow[0], ColDict["Booster"]+2).value = BoosterInfo.SWVersion
                        MainInfoSheet.cell(VINRow[0], ColDict["Booster"]+3).value = BoosterInfo.KeyType

            #Get DB Data
            # SubItemIter = GetIterator(temp, "SubTeilnehmer") 
            if SubItemIter != None:   
                DBtemp = SubItemIter.getchildren()  
                for DBTemp in DBtemp:
                    DBChild = DBTemp.getchildren()  
                    DBIter = GetIterator(DBChild, "SubtName")
                    # print (DBIter.text)
                    if DBIter.text == "Data Medium":
                        GetInfoData(DBChild, DBInfo)
                        MainInfoSheet.cell(VINRow[0], ColDict["DB"]).value = DBInfo.DBINFO

    
        elif InfoIter.text == "Telematics Communication Unit": #Get Cbox Data
            GetInfoData(temp, cBoxInfo)       
            MainInfoSheet.cell(VINRow[0], ColDict["CBOX/OCU"]).value = cBoxInfo.PartNo
            MainInfoSheet.cell(VINRow[0], ColDict["CBOX/OCU"]+1).value = cBoxInfo.HWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["CBOX/OCU"]+2).value = cBoxInfo.SWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["CBOX/OCU"]+3).value = cBoxInfo.KeyType 

        elif InfoIter.text == "Gateway": #Get Gateway Data
            GetInfoData(temp, GWInfo)        
            MainInfoSheet.cell(VINRow[0], ColDict["Gateway"]).value = GWInfo.PartNo
            MainInfoSheet.cell(VINRow[0], ColDict["Gateway"]+1).value = GWInfo.HWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Gateway"]+2).value = GWInfo.SWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Gateway"]+3).value = GWInfo.KeyType
        
        elif InfoIter.text == "Dash Board": #Get Kombi Data
            GetInfoData(temp, KombiInfo)    
           
            MainInfoSheet.cell(VINRow[0], ColDict["Kombi"]).value = KombiInfo.PartNo
            MainInfoSheet.cell(VINRow[0], ColDict["Kombi"]+1).value = KombiInfo.HWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Kombi"]+2).value = KombiInfo.SWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Kombi"]+3).value = KombiInfo.KeyType            

        elif InfoIter.text == "Central Electrics": #Get BCM1 Data
            GetInfoData(temp, BCM1Info)  

            MainInfoSheet.cell(VINRow[0], ColDict["BCM1"]).value = BCM1Info.PartNo
            MainInfoSheet.cell(VINRow[0], ColDict["BCM1"]+1).value = BCM1Info.HWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["BCM1"]+2).value = BCM1Info.SWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["BCM1"]+3).value = BCM1Info.KeyType
      
        elif InfoIter.text == "Central Module Comfort System": #Get BCM2 Data
            GetInfoData(temp, BCM2Info)

            MainInfoSheet.cell(VINRow[0], ColDict["BCM2"]).value = BCM2Info.PartNo
            MainInfoSheet.cell(VINRow[0], ColDict["BCM2"]+1).value = BCM2Info.HWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["BCM2"]+2).value = BCM2Info.SWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["BCM2"]+3).value = BCM2Info.KeyType
        
        elif InfoIter.text == "Head Up Display": #Get HUD Data
            GetInfoData(temp, HUDInfo)

            MainInfoSheet.cell(VINRow[0], ColDict["Head Up Display"]).value = HUDInfo.PartNo
            MainInfoSheet.cell(VINRow[0], ColDict["Head Up Display"]+1).value = HUDInfo.HWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Head Up Display"]+2).value = HUDInfo.SWVersion
            MainInfoSheet.cell(VINRow[0], ColDict["Head Up Display"]+3).value = HUDInfo.KeyType
    
    print("Import: ", xmlIndex)

    MainInfoSheet.cell(VINRow[0], ColDict["Updated Date"]).value = time.asctime()
    #Write to Excel
fExcelFile.save(filename = sExcelPath) #Save Excel File

print("Successfully done. Please press any key.")
End = input()
# CurTime = str(datetime.now())
# CurTime = CurTime.replace("-","_")
# CurTime = CurTime.replace(":","_")
# CurTime = CurTime.replace(".","_")
# sBackupExcelPath = "\\\10.104.97.225\\mmi\\99_Test_Car_Bench_Status (hard & software documentation)\\00_Maintenance\\History\\AVK Test Environment Status_" + CurTime + ".xlsx"
# # sBackupExcelPath = "D:\\24_Python_Dev\\AVK Test Environment Status_" + CurTime + ".xlsx"
# fExcelFile.save(filename = sBackupExcelPath) #Save Excel File




    
    

